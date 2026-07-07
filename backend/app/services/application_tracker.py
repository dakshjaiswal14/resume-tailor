"""
Application tracker — stores job applications in a local Excel file.
"""

import os
import uuid
from datetime import datetime, timezone
from pathlib import Path
from typing import Optional

import openpyxl

ROOT_DIR = Path(__file__).resolve().parents[3]
EXCEL_PATH = ROOT_DIR / "applications.xlsx"

COLUMNS = [
    "id", "company", "position", "date_applied", "status",
    "resume_id", "cover_letter_text", "notes", "last_updated",
]
STATUSES = ["Applied", "Phone Screen", "Technical", "Onsite", "Offer", "Accepted", "Rejected", "Withdrawn"]


def _ensure_sheet() -> openpyxl.Workbook:
    """Load or create the Excel workbook with headers."""
    if EXCEL_PATH.exists():
        wb = openpyxl.load_workbook(EXCEL_PATH)
        if "Applications" not in wb.sheetnames:
            ws = wb.active
            ws.title = "Applications"
        else:
            ws = wb["Applications"]
            # Ensure headers
            if ws.max_row == 0 or ws.cell(1, 1).value != "id":
                ws.insert_rows(1)
                for col_idx, header in enumerate(COLUMNS, 1):
                    ws.cell(1, col_idx, header)
        wb.save(EXCEL_PATH)
        return wb

    # Create new
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Applications"
    for col_idx, header in enumerate(COLUMNS, 1):
        ws.cell(1, col_idx, header)
        ws.cell(1, col_idx).font = openpyxl.styles.Font(bold=True)
    wb.save(EXCEL_PATH)
    return wb


def _row_to_dict(ws, row_idx: int) -> dict:
    """Convert a worksheet row to a dict."""
    return {
        COLUMNS[i]: ws.cell(row_idx, i + 1).value
        for i in range(len(COLUMNS))
    }


def list_applications() -> list[dict]:
    """Return all applications sorted by date (newest first)."""
    wb = _ensure_sheet()
    ws = wb["Applications"]
    rows = []
    for row_idx in range(2, ws.max_row + 1):
        app = _row_to_dict(ws, row_idx)
        if app["id"]:  # skip empty rows
            rows.append(app)
    wb.close()
    # Sort by date_applied descending
    rows.sort(key=lambda r: str(r.get("date_applied") or ""), reverse=True)
    return rows


def add_application(
    company: str,
    position: str,
    resume_id: str = "",
    cover_letter_text: str = "",
    notes: str = "",
) -> dict:
    """Add a new application and return it."""
    wb = _ensure_sheet()
    ws = wb["Applications"]

    now = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M")
    app = {
        "id": uuid.uuid4().hex[:8],
        "company": company,
        "position": position,
        "date_applied": datetime.now().strftime("%Y-%m-%d"),
        "status": "Applied",
        "resume_id": resume_id,
        "cover_letter_text": cover_letter_text[:500] if cover_letter_text else "",  # truncate for Excel
        "notes": notes,
        "last_updated": now,
    }

    row_idx = ws.max_row + 1
    for col_idx, key in enumerate(COLUMNS, 1):
        ws.cell(row_idx, col_idx, app[key])

    wb.save(EXCEL_PATH)
    wb.close()
    return app


def update_application(app_id: str, status: str = "", notes: str = "") -> Optional[dict]:
    """Update the status and/or notes of an application."""
    if status and status not in STATUSES:
        raise ValueError(f"Invalid status: {status}. Must be one of {STATUSES}")

    wb = _ensure_sheet()
    ws = wb["Applications"]

    for row_idx in range(2, ws.max_row + 1):
        if ws.cell(row_idx, 1).value == app_id:
            if status:
                ws.cell(row_idx, 5, status)  # column 5 = status
            if notes:
                existing = ws.cell(row_idx, 8).value or ""
                ws.cell(row_idx, 8, existing + ("\n" if existing else "") + notes)
            ws.cell(row_idx, 9, datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M"))
            wb.save(EXCEL_PATH)
            wb.close()
            return _row_to_dict(ws, row_idx)

    wb.close()
    return None


def delete_application(app_id: str) -> bool:
    """Delete an application by ID. Returns True if deleted."""
    wb = _ensure_sheet()
    ws = wb["Applications"]

    for row_idx in range(2, ws.max_row + 1):
        if ws.cell(row_idx, 1).value == app_id:
            ws.delete_rows(row_idx)
            wb.save(EXCEL_PATH)
            wb.close()
            return True

    wb.close()
    return False
